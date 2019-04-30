import random
import math
import string
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from PIL import Image,ImageEnhance
from openpyxl import Workbook

def suffixGenerator(size=5, chars=string.ascii_uppercase + string.digits):
  return ''.join(random.choice(chars) for _ in range(size))

def signUpChinesegamerAccount(ChinesegamerAccount, Password):
    driver = webdriver.Chrome()
    driver.set_window_position(-3000, 0)
    url = r"https://www.chinesegamer.net/check1.asp?checkbox=on&checkbox2=on&Submit=%A4U%A4%40%A8B"
    driver.get(url)
    accountInput = driver.find_element_by_name("a2")
    accountInput.send_keys(ChinesegamerAccount)
    passwordInput = driver.find_element_by_name("a3a")
    passwordInput.send_keys(Password)
    confrimPasswordInput = driver.find_element_by_name("a3b")
    confrimPasswordInput.send_keys(Password)
    getCaptchaButton = driver.find_element_by_xpath("//img[contains(@src,'/images/code.gif')]")
    getCaptchaButton.click()
    driver.get_screenshot_as_file("captcha.png")
    location = driver.find_element_by_name('vcjpg').location
    size = driver.find_element_by_name('vcjpg').size
    left = location['x']
    top =  location['y']
    right = location['x'] + size['width']
    bottom = location['y'] + size['height']
    img = Image.open("captcha.png").crop((left,top,right,bottom))
    img = img.convert('L') 		
    img = ImageEnhance.Contrast(img)
    img = img.enhance(5.0) 	
    img = img.resize((100,50))
    img.show()
    captcha = input('Please Input Captcha: ')
    driver.find_element_by_name("a4").send_keys(captcha)
    driver.find_element_by_name("Submit").click()
    try:
      driver.find_element_by_name("a4").send_keys('asdasdas')
      s1 = Select(driver.find_element_by_name('a24y'))
      s1.select_by_index(1)
      s2 = Select(driver.find_element_by_name('a24m'))
      s2.select_by_index(1)
      driver.find_element_by_name("a25x").send_keys('12345')
      s3 = Select(driver.find_element_by_name('a6'))
      s3.select_by_index(1)
      driver.find_element_by_name("a8").send_keys('85256985214')
      driver.find_element_by_name("Submit").click()
    except NoSuchElementException:
      driver.quit()
      print('Somthing error cannot no redirect to next step. trying again ( maybe wrong captcha )')
      signUpChinesegamerAccount(ChinesegamerAccount, Password)
    if driver.find_element_by_class_name("f1").text == '您的帳號已註冊成功，恭喜您成為中華遊戲網會員。':
      driver.quit()
      return True

def signUptspAccount(ChinesegamerAccount, Password):
  result = []
  driver = webdriver.Chrome()
  driver.set_window_position(-3000, 0)
  url = r"https://tspac.chinesegamer.net/"
  driver.get(url)
  accountInput = driver.find_element_by_name("txtAccount")
  accountInput.send_keys(ChinesegamerAccount)
  passwordInput = driver.find_element_by_name("txtPassword")
  passwordInput.send_keys(Password)
  driver.find_element_by_name("Button1").click()
  try:
    result.append(driver.find_element_by_id("lblAccMessage").text)
    result.append(driver.find_element_by_id("lblPassMessage").text)
    driver.quit()
    return result
  except NoSuchElementException:
    return False

def signUpStart(TotalOfAccounts, AccountsPrefix, Password):
    TotalOfRegisteredAccount = 0
    wb = Workbook()
    ws = wb.active  
    for TotalOfRegisteredChinesegamerAccount in range(int(math.ceil(TotalOfAccounts / 3))):
      ChinesegamerAccount =  AccountsPrefix + suffixGenerator()
      print('Now try to register chinesegamer account by {}'.format(ChinesegamerAccount))
      if signUpChinesegamerAccount(ChinesegamerAccount, Password):
        print('Try to register chinesegamer account by {} success'.format(ChinesegamerAccount))
        print('Starting to register tsp account')
        for TotalOfRegisteredGameAccount in range(3):
          result = signUptspAccount(ChinesegamerAccount, Password)
          print('TSP account registed account: {}, password: {} .'.format(result[0],result[1]))
          ws.cell(row=TotalOfRegisteredAccount + 1, column=1, value=str(result[0]))
          ws.cell(row=TotalOfRegisteredAccount + 1, column=2, value=str(result[1]))
          ws.cell(row=TotalOfRegisteredAccount + 1, column=3, value=str(ChinesegamerAccount))
          ws.cell(row=TotalOfRegisteredAccount + 1, column=4, value=str(Password))
          TotalOfRegisteredAccount += 1
      else:
        print('Something error, exit script')
        exit()
    wb.save('result.xlsx')

    return True
    
TotalOfAccounts = input('How many TSP accounts do you want? ( It should be multiples of 3) ')
AccountsPrefix = input('What prefix of Chinesegamer do you want? ( Only support english ) ')
Password = input('Please input password of Chinesegamer do you want ( Only support english ) ')
print('Since Chinesegamer account name cannot be repeated, 5 random english and number will become your account suffix.')
yes = {'yes','y', 'ye', ''}
Confrim = input('Did you want to start now? (yes / no)').lower()
if Confrim in yes:
  signUpStart(int(TotalOfAccounts), AccountsPrefix, Password)
else:
  print('Okay, Bye')
  exit()
